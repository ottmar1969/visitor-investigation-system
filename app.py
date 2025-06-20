"""
Enhanced Automated Client Onboarding with Trial Management
Self-service client registration with flexible trial periods and automatic restrictions
"""

from flask import Flask, render_template, request, jsonify, redirect, url_for, session, send_file
from flask_cors import CORS
import sqlite3
import json
import random
import string
import hashlib
import time
import uuid
from datetime import datetime, timedelta
import secrets
import os
import csv
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import requests
import smtplib
from email.mime.text import MimeText
from email.mime.multipart import MimeMultipart
import threading
import schedule

app = Flask(__name__)
app.secret_key = secrets.token_hex(32)
CORS(app, origins="*")

# Email configuration (configure with your SMTP settings)
EMAIL_CONFIG = {
    'smtp_server': 'smtp.gmail.com',  # Change to your SMTP server
    'smtp_port': 587,
    'email': 'your-email@gmail.com',  # Change to your email
    'password': 'your-app-password',  # Change to your app password
    'from_name': 'Visitor Investigation System'
}

def init_database():
    """Initialize database with trial management tables"""
    conn = sqlite3.connect('client_management.db')
    cursor = conn.cursor()
    
    # Enhanced clients table with trial support
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS clients (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            client_id TEXT UNIQUE NOT NULL,
            business_name TEXT NOT NULL,
            contact_email TEXT NOT NULL,
            website_url TEXT NOT NULL,
            access_token TEXT UNIQUE NOT NULL,
            subscription_status TEXT DEFAULT 'active',
            account_type TEXT DEFAULT 'full',
            trial_start_time TIMESTAMP,
            trial_end_time TIMESTAMP,
            trial_duration_hours INTEGER,
            trial_extended_count INTEGER DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            last_access TIMESTAMP,
            billing_cycle TEXT DEFAULT 'monthly',
            plan_type TEXT DEFAULT 'basic',
            max_users INTEGER DEFAULT 5,
            owner_user_id TEXT,
            auto_restricted_at TIMESTAMP,
            conversion_date TIMESTAMP,
            trial_usage_stats TEXT
        )
    ''')
    
    # Trial management table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS trial_management (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            trial_id TEXT UNIQUE NOT NULL,
            client_id TEXT NOT NULL,
            granted_by TEXT,
            trial_type TEXT DEFAULT 'standard',
            duration_hours INTEGER NOT NULL,
            start_time TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            end_time TIMESTAMP NOT NULL,
            status TEXT DEFAULT 'active',
            usage_stats TEXT,
            reminder_sent BOOLEAN DEFAULT 0,
            expiration_warning_sent BOOLEAN DEFAULT 0,
            auto_restricted_at TIMESTAMP,
            extension_count INTEGER DEFAULT 0,
            conversion_attempted BOOLEAN DEFAULT 0,
            notes TEXT,
            FOREIGN KEY (client_id) REFERENCES clients (client_id)
        )
    ''')
    
    # Trial notifications table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS trial_notifications (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            notification_id TEXT UNIQUE NOT NULL,
            client_id TEXT NOT NULL,
            notification_type TEXT NOT NULL,
            scheduled_time TIMESTAMP NOT NULL,
            sent_time TIMESTAMP,
            status TEXT DEFAULT 'pending',
            email_content TEXT,
            retry_count INTEGER DEFAULT 0,
            FOREIGN KEY (client_id) REFERENCES clients (client_id)
        )
    ''')
    
    # Enhanced client users table with trial restrictions
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS client_users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id TEXT UNIQUE NOT NULL,
            client_id TEXT NOT NULL,
            name TEXT NOT NULL,
            email TEXT NOT NULL,
            role TEXT DEFAULT 'viewer',
            access_token TEXT UNIQUE NOT NULL,
            status TEXT DEFAULT 'active',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            created_by TEXT,
            last_access TIMESTAMP,
            access_expires_at TIMESTAMP,
            trial_restricted BOOLEAN DEFAULT 0,
            country_restrictions TEXT,
            block_vpn BOOLEAN DEFAULT 0,
            permissions TEXT,
            session_limit INTEGER DEFAULT 1,
            current_sessions INTEGER DEFAULT 0,
            notes TEXT,
            FOREIGN KEY (client_id) REFERENCES clients (client_id)
        )
    ''')
    
    # Visitor investigations table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS visitor_investigations (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            visitor_id TEXT UNIQUE NOT NULL,
            client_id TEXT NOT NULL,
            ip_address TEXT NOT NULL,
            user_agent TEXT,
            country TEXT,
            city TEXT,
            device_type TEXT,
            browser TEXT,
            operating_system TEXT,
            referrer TEXT,
            landing_page TEXT,
            current_page TEXT,
            pages_visited TEXT,
            time_on_site INTEGER,
            session_start TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            last_activity TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            interest_level TEXT DEFAULT 'medium',
            contact_info TEXT,
            company_info TEXT,
            notes TEXT,
            status TEXT DEFAULT 'active',
            FOREIGN KEY (client_id) REFERENCES clients (client_id)
        )
    ''')
    
    # User sessions table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS user_sessions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            session_id TEXT UNIQUE NOT NULL,
            user_id TEXT NOT NULL,
            client_id TEXT NOT NULL,
            ip_address TEXT,
            user_agent TEXT,
            country TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            last_activity TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            is_active BOOLEAN DEFAULT 1,
            FOREIGN KEY (user_id) REFERENCES client_users (user_id),
            FOREIGN KEY (client_id) REFERENCES clients (client_id)
        )
    ''')
    
    # Access logs table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS access_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            log_id TEXT UNIQUE NOT NULL,
            user_id TEXT,
            client_id TEXT,
            action TEXT NOT NULL,
            resource TEXT,
            ip_address TEXT,
            user_agent TEXT,
            country TEXT,
            timestamp TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            success BOOLEAN DEFAULT 1,
            error_message TEXT,
            FOREIGN KEY (user_id) REFERENCES client_users (user_id),
            FOREIGN KEY (client_id) REFERENCES clients (client_id)
        )
    ''')
    
    # Automated tasks table with trial management
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS automated_tasks (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            task_id TEXT UNIQUE NOT NULL,
            task_type TEXT NOT NULL,
            client_id TEXT,
            trial_id TEXT,
            status TEXT DEFAULT 'pending',
            scheduled_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            executed_at TIMESTAMP,
            result TEXT,
            error_message TEXT,
            retry_count INTEGER DEFAULT 0,
            max_retries INTEGER DEFAULT 3,
            task_data TEXT
        )
    ''')
    
    conn.commit()
    conn.close()

# Initialize database
init_database()

def generate_client_id():
    """Generate unique client identifier"""
    return f"client_{uuid.uuid4().hex[:12]}"

def generate_user_id():
    """Generate unique user identifier"""
    return f"user_{uuid.uuid4().hex[:12]}"

def generate_access_token():
    """Generate secure access token"""
    return secrets.token_urlsafe(32)

def generate_trial_id():
    """Generate unique trial identifier"""
    return f"trial_{uuid.uuid4().hex[:12]}"

def calculate_trial_end_time(duration_hours):
    """Calculate trial end time based on duration"""
    return datetime.now() + timedelta(hours=duration_hours)

def create_trial_account(business_name, contact_email, website_url, duration_hours, granted_by='admin', trial_type='standard'):
    """Create a trial account with automatic expiration"""
    try:
        client_id = generate_client_id()
        trial_id = generate_trial_id()
        owner_user_id = generate_user_id()
        client_access_token = generate_access_token()
        owner_access_token = generate_access_token()
        
        trial_start = datetime.now()
        trial_end = calculate_trial_end_time(duration_hours)
        
        conn = sqlite3.connect('client_management.db')
        cursor = conn.cursor()
        
        # Create trial client
        cursor.execute('''
            INSERT INTO clients 
            (client_id, business_name, contact_email, website_url, access_token, 
             account_type, trial_start_time, trial_end_time, trial_duration_hours,
             plan_type, max_users, owner_user_id, subscription_status)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (client_id, business_name, contact_email, website_url, client_access_token, 
              'trial', trial_start.isoformat(), trial_end.isoformat(), duration_hours,
              'trial', 3, owner_user_id, 'trial'))  # Trial accounts get 3 users max
        
        # Create trial management record
        cursor.execute('''
            INSERT INTO trial_management 
            (trial_id, client_id, granted_by, trial_type, duration_hours, 
             start_time, end_time, usage_stats)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        ''', (trial_id, client_id, granted_by, trial_type, duration_hours,
              trial_start.isoformat(), trial_end.isoformat(), json.dumps({})))
        
        # Create owner user with trial permissions
        trial_permissions = {
            'view_visitors': True,
            'view_contact_info': True,
            'view_email': True,
            'view_phone': False,  # Limited in trial
            'view_company': True,
            'export_data': False,  # Limited in trial
            'manage_users': True,
            'view_audit_logs': False  # Limited in trial
        }
        
        cursor.execute('''
            INSERT INTO client_users 
            (user_id, client_id, name, email, role, access_token, created_by, 
             permissions, session_limit)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (owner_user_id, client_id, business_name + ' (Trial)', contact_email, 'owner', 
              owner_access_token, granted_by, json.dumps(trial_permissions), 2))
        
        # Schedule trial notifications
        schedule_trial_notifications(client_id, trial_id, trial_end, duration_hours)
        
        # Schedule automatic restriction
        schedule_automatic_restriction(client_id, trial_id, trial_end)
        
        conn.commit()
        conn.close()
        
        # Generate demo data for trial
        create_automated_task('generate_demo_data', client_id, trial_id)
        
        return {
            'success': True,
            'client_id': client_id,
            'trial_id': trial_id,
            'dashboard_url': f"http://localhost:5000/dashboard/{owner_access_token}",
            'access_token': owner_access_token,
            'trial_end_time': trial_end.isoformat(),
            'duration_hours': duration_hours
        }
        
    except Exception as e:
        print(f"Error creating trial account: {e}")
        return {'success': False, 'error': str(e)}

def schedule_trial_notifications(client_id, trial_id, trial_end, duration_hours):
    """Schedule trial reminder and expiration notifications"""
    conn = sqlite3.connect('client_management.db')
    cursor = conn.cursor()
    
    # Schedule reminder notification (at 75% of trial period)
    reminder_time = trial_end - timedelta(hours=duration_hours * 0.25)
    if reminder_time > datetime.now():
        cursor.execute('''
            INSERT INTO trial_notifications 
            (notification_id, client_id, notification_type, scheduled_time)
            VALUES (?, ?, ?, ?)
        ''', (f"reminder_{uuid.uuid4().hex[:8]}", client_id, 'trial_reminder', reminder_time.isoformat()))
    
    # Schedule expiration warning (1 hour before expiration for short trials, 24 hours for long trials)
    warning_hours = 1 if duration_hours <= 24 else 24
    warning_time = trial_end - timedelta(hours=warning_hours)
    if warning_time > datetime.now():
        cursor.execute('''
            INSERT INTO trial_notifications 
            (notification_id, client_id, notification_type, scheduled_time)
            VALUES (?, ?, ?, ?)
        ''', (f"warning_{uuid.uuid4().hex[:8]}", client_id, 'trial_expiring', warning_time.isoformat()))
    
    # Schedule expiration notification (at trial end)
    cursor.execute('''
        INSERT INTO trial_notifications 
        (notification_id, client_id, notification_type, scheduled_time)
        VALUES (?, ?, ?, ?)
    ''', (f"expired_{uuid.uuid4().hex[:8]}", client_id, 'trial_expired', trial_end.isoformat()))
    
    conn.commit()
    conn.close()

def schedule_automatic_restriction(client_id, trial_id, trial_end):
    """Schedule automatic restriction task"""
    create_automated_task('restrict_trial_access', client_id, trial_id, trial_end)

def check_and_restrict_expired_trials():
    """Check for expired trials and automatically restrict access"""
    conn = sqlite3.connect('client_management.db')
    cursor = conn.cursor()
    
    # Find expired trials that haven't been restricted yet
    cursor.execute('''
        SELECT client_id, trial_id, business_name, contact_email
        FROM clients c
        JOIN trial_management tm ON c.client_id = tm.client_id
        WHERE c.account_type = 'trial' 
        AND c.trial_end_time <= ? 
        AND c.auto_restricted_at IS NULL
        AND tm.status = 'active'
    ''', (datetime.now().isoformat(),))
    
    expired_trials = cursor.fetchall()
    
    for client_id, trial_id, business_name, contact_email in expired_trials:
        try:
            # Restrict client access
            cursor.execute('''
                UPDATE clients 
                SET subscription_status = 'trial_expired', auto_restricted_at = ?
                WHERE client_id = ?
            ''', (datetime.now().isoformat(), client_id))
            
            # Restrict all users for this client
            cursor.execute('''
                UPDATE client_users 
                SET status = 'trial_expired', trial_restricted = 1
                WHERE client_id = ?
            ''', (client_id,))
            
            # Deactivate all sessions
            cursor.execute('''
                UPDATE user_sessions 
                SET is_active = 0 
                WHERE client_id = ?
            ''', (client_id,))
            
            # Update trial management
            cursor.execute('''
                UPDATE trial_management 
                SET status = 'expired', auto_restricted_at = ?
                WHERE trial_id = ?
            ''', (datetime.now().isoformat(), trial_id))
            
            print(f"Automatically restricted expired trial for {business_name}")
            
        except Exception as e:
            print(f"Error restricting trial {trial_id}: {e}")
    
    conn.commit()
    conn.close()
    
    return len(expired_trials)

def extend_trial(client_id, additional_hours, extended_by='admin'):
    """Extend an existing trial"""
    try:
        conn = sqlite3.connect('client_management.db')
        cursor = conn.cursor()
        
        # Get current trial info
        cursor.execute('''
            SELECT trial_end_time, trial_extended_count
            FROM clients 
            WHERE client_id = ? AND account_type = 'trial'
        ''', (client_id,))
        
        result = cursor.fetchone()
        if not result:
            return {'success': False, 'error': 'Trial not found'}
        
        current_end_time, extension_count = result
        current_end = datetime.fromisoformat(current_end_time)
        new_end_time = current_end + timedelta(hours=additional_hours)
        
        # Update client trial end time
        cursor.execute('''
            UPDATE clients 
            SET trial_end_time = ?, trial_extended_count = ?,
                subscription_status = 'trial', auto_restricted_at = NULL
            WHERE client_id = ?
        ''', (new_end_time.isoformat(), extension_count + 1, client_id))
        
        # Reactivate users if they were restricted
        cursor.execute('''
            UPDATE client_users 
            SET status = 'active', trial_restricted = 0
            WHERE client_id = ? AND trial_restricted = 1
        ''', (client_id,))
        
        # Update trial management
        cursor.execute('''
            UPDATE trial_management 
            SET end_time = ?, extension_count = ?, status = 'active'
            WHERE client_id = ?
        ''', (new_end_time.isoformat(), extension_count + 1, client_id))
        
        # Schedule new notifications
        cursor.execute('SELECT trial_id FROM trial_management WHERE client_id = ?', (client_id,))
        trial_id = cursor.fetchone()[0]
        
        schedule_trial_notifications(client_id, trial_id, new_end_time, additional_hours)
        schedule_automatic_restriction(client_id, trial_id, new_end_time)
        
        conn.commit()
        conn.close()
        
        return {
            'success': True,
            'new_end_time': new_end_time.isoformat(),
            'extension_count': extension_count + 1
        }
        
    except Exception as e:
        return {'success': False, 'error': str(e)}

def convert_trial_to_full(client_id, plan_type='basic'):
    """Convert trial account to full paid account"""
    try:
        conn = sqlite3.connect('client_management.db')
        cursor = conn.cursor()
        
        # Update client to full account
        max_users = {'basic': 5, 'professional': 15, 'enterprise': 50}.get(plan_type, 5)
        
        cursor.execute('''
            UPDATE clients 
            SET account_type = 'full', subscription_status = 'active', 
                plan_type = ?, max_users = ?, conversion_date = ?
            WHERE client_id = ?
        ''', (plan_type, max_users, datetime.now().isoformat(), client_id))
        
        # Update trial management
        cursor.execute('''
            UPDATE trial_management 
            SET status = 'converted', conversion_attempted = 1
            WHERE client_id = ?
        ''', (client_id,))
        
        # Upgrade user permissions to full access
        full_permissions = {
            'view_visitors': True,
            'view_contact_info': True,
            'view_email': True,
            'view_phone': True,
            'view_company': True,
            'export_data': True,
            'manage_users': True,
            'view_audit_logs': True
        }
        
        cursor.execute('''
            UPDATE client_users 
            SET permissions = ?, status = 'active', trial_restricted = 0
            WHERE client_id = ? AND role = 'owner'
        ''', (json.dumps(full_permissions), client_id))
        
        conn.commit()
        conn.close()
        
        return {'success': True, 'message': 'Trial converted to full account successfully'}
        
    except Exception as e:
        return {'success': False, 'error': str(e)}

def create_automated_task(task_type, client_id=None, trial_id=None, scheduled_at=None, task_data=None):
    """Create an automated task with trial support"""
    task_id = f"task_{uuid.uuid4().hex[:12]}"
    if not scheduled_at:
        scheduled_at = datetime.now()
    
    conn = sqlite3.connect('client_management.db')
    cursor = conn.cursor()
    
    cursor.execute('''
        INSERT INTO automated_tasks (task_id, task_type, client_id, trial_id, scheduled_at, task_data)
        VALUES (?, ?, ?, ?, ?, ?)
    ''', (task_id, task_type, client_id, trial_id, scheduled_at.isoformat(), 
          json.dumps(task_data) if task_data else None))
    
    conn.commit()
    conn.close()
    
    return task_id

def generate_demo_visitors(client_id, count=75):
    """Generate demo visitor data for trials"""
    conn = sqlite3.connect('client_management.db')
    cursor = conn.cursor()
    
    # Sample data for realistic demo
    companies = ['Acme Corp', 'TechStart Inc', 'Global Solutions', 'Innovation Labs', 'Digital Dynamics']
    pages = ['/home', '/about', '/services', '/contact', '/pricing', '/blog', '/products', '/support']
    countries = ['United States', 'Canada', 'United Kingdom', 'Germany', 'Australia']
    cities = ['New York', 'Toronto', 'London', 'Berlin', 'Sydney']
    browsers = ['Chrome', 'Firefox', 'Safari', 'Edge']
    devices = ['Desktop', 'Mobile', 'Tablet']
    
    for i in range(count):
        visitor_id = f"demo_{uuid.uuid4().hex[:12]}"
        
        # Generate realistic visitor data
        pages_visited = random.sample(pages, random.randint(1, min(len(pages), 25)))  # Up to 25 pages
        time_on_site = random.randint(30, 3600)  # 30 seconds to 1 hour
        interest_levels = ['high', 'medium', 'low']
        weights = [0.2, 0.5, 0.3]  # 20% high, 50% medium, 30% low
        interest_level = random.choices(interest_levels, weights=weights)[0]
        
        # Generate contact info for some visitors
        contact_info = None
        if random.random() < 0.3:  # 30% have contact info
            contact_info = json.dumps({
                'email': f'visitor{i}@{random.choice(companies).lower().replace(" ", "")}.com',
                'phone': f'+1-555-{random.randint(100, 999)}-{random.randint(1000, 9999)}',
                'name': f'Visitor {i+1}'
            })
        
        cursor.execute('''
            INSERT INTO visitor_investigations 
            (visitor_id, client_id, ip_address, user_agent, country, city, 
             device_type, browser, operating_system, referrer, landing_page, 
             current_page, pages_visited, time_on_site, interest_level, 
             contact_info, company_info, status)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            visitor_id, client_id, 
            f"192.168.{random.randint(1, 255)}.{random.randint(1, 255)}",
            f"Mozilla/5.0 ({random.choice(['Windows NT 10.0', 'Macintosh', 'X11; Linux x86_64'])})",
            random.choice(countries), random.choice(cities),
            random.choice(devices), random.choice(browsers), 
            random.choice(['Windows', 'macOS', 'Linux', 'iOS', 'Android']),
            random.choice(['Google', 'Direct', 'Social Media', 'Email']),
            pages_visited[0], pages_visited[-1],
            json.dumps(pages_visited), time_on_site, interest_level,
            contact_info, 
            json.dumps({'company': random.choice(companies)}) if random.random() < 0.4 else None,
            'active'
        ))
    
    conn.commit()
    conn.close()

def verify_user_access_with_trial(access_token, ip_address=None, user_agent=None):
    """Verify user access with trial expiration checking"""
    conn = sqlite3.connect('client_management.db')
    cursor = conn.cursor()
    
    # Get user and client info
    cursor.execute('''
        SELECT cu.user_id, cu.client_id, cu.name, cu.email, cu.role, cu.status,
               cu.permissions, cu.trial_restricted, c.account_type, c.trial_end_time,
               c.subscription_status, c.business_name
        FROM client_users cu
        JOIN clients c ON cu.client_id = c.client_id
        WHERE cu.access_token = ? AND cu.status = 'active'
    ''', (access_token,))
    
    result = cursor.fetchone()
    
    if not result:
        conn.close()
        return None
    
    user_id, client_id, name, email, role, status, permissions, trial_restricted, account_type, trial_end_time, subscription_status, business_name = result
    
    # Check if trial has expired
    if account_type == 'trial' and trial_end_time:
        trial_end = datetime.fromisoformat(trial_end_time)
        if datetime.now() >= trial_end:
            # Trial has expired, restrict access
            cursor.execute('''
                UPDATE client_users 
                SET status = 'trial_expired', trial_restricted = 1
                WHERE user_id = ?
            ''', (user_id,))
            conn.commit()
            conn.close()
            return None
    
    # Parse permissions
    try:
        user_permissions = json.loads(permissions) if permissions else {}
    except:
        user_permissions = {}
    
    # Update last access
    cursor.execute('''
        UPDATE client_users 
        SET last_access = ?
        WHERE user_id = ?
    ''', (datetime.now().isoformat(), user_id))
    
    conn.commit()
    conn.close()
    
    user_data = {
        'user_id': user_id,
        'client_id': client_id,
        'name': name,
        'email': email,
        'role': role,
        'permissions': user_permissions,
        'account_type': account_type,
        'business_name': business_name,
        'is_trial': account_type == 'trial'
    }
    
    if account_type == 'trial' and trial_end_time:
        trial_end = datetime.fromisoformat(trial_end_time)
        time_remaining = trial_end - datetime.now()
        user_data['trial_end_time'] = trial_end_time
        user_data['trial_hours_remaining'] = max(0, int(time_remaining.total_seconds() / 3600))
    
    return user_data

# Flask Routes

@app.route('/')
def index():
    """Main visitor investigation page"""
    return render_template('index.html')

@app.route('/dashboard/<access_token>')
def client_dashboard(access_token):
    """Client dashboard with visitor data"""
    user_data = verify_user_access_with_trial(access_token)
    
    if not user_data:
        return render_template('access_denied.html'), 403
    
    return render_template('client_dashboard.html', user_data=user_data)

@app.route('/admin')
def admin_dashboard():
    """Admin dashboard for managing clients"""
    return render_template('admin_dashboard.html')

@app.route('/admin/trials')
def admin_trials():
    """Admin interface for trial management"""
    return render_template('admin_trials.html')

@app.route('/admin/users')
def admin_users():
    """Admin interface for user management"""
    return render_template('user_management.html')

# API Routes

@app.route('/api/create-trial', methods=['POST'])
def create_trial():
    """Create a new trial account"""
    try:
        data = request.get_json()
        business_name = data.get('business_name', '').strip()
        contact_email = data.get('contact_email', '').strip()
        website_url = data.get('website_url', '').strip()
        duration_hours = int(data.get('duration_hours', 24))
        trial_type = data.get('trial_type', 'standard')
        granted_by = data.get('granted_by', 'admin')
        
        if not all([business_name, contact_email, website_url]):
            return jsonify({'error': 'Business name, email, and website URL are required'}), 400
        
        if duration_hours < 1 or duration_hours > 8760:  # Max 1 year
            return jsonify({'error': 'Duration must be between 1 hour and 8760 hours (1 year)'}), 400
        
        result = create_trial_account(business_name, contact_email, website_url, 
                                    duration_hours, granted_by, trial_type)
        
        if result['success']:
            # Generate demo data
            generate_demo_visitors(result['client_id'])
            return jsonify(result)
        else:
            return jsonify({'error': result['error']}), 500
            
    except Exception as e:
        return jsonify({'error': f'Failed to create trial: {str(e)}'}), 500

@app.route('/api/trials')
def get_trials():
    """Get all trial accounts"""
    conn = sqlite3.connect('client_management.db')
    cursor = conn.cursor()
    
    cursor.execute('''
        SELECT c.client_id, c.business_name, c.contact_email, c.website_url,
               c.trial_start_time, c.trial_end_time, c.trial_duration_hours,
               c.trial_extended_count, c.subscription_status, c.auto_restricted_at,
               c.conversion_date, tm.trial_id, tm.granted_by, tm.trial_type,
               tm.status, tm.extension_count
        FROM clients c
        JOIN trial_management tm ON c.client_id = tm.client_id
        WHERE c.account_type = 'trial'
        ORDER BY c.created_at DESC
    ''')
    
    trials = []
    for row in cursor.fetchall():
        trial_end = datetime.fromisoformat(row[5]) if row[5] else None
        time_remaining = None
        if trial_end and datetime.now() < trial_end:
            time_remaining = int((trial_end - datetime.now()).total_seconds() / 3600)  # Hours remaining
        
        trials.append({
            'client_id': row[0],
            'business_name': row[1],
            'contact_email': row[2],
            'website_url': row[3],
            'trial_start_time': row[4],
            'trial_end_time': row[5],
            'trial_duration_hours': row[6],
            'trial_extended_count': row[7],
            'subscription_status': row[8],
            'auto_restricted_at': row[9],
            'conversion_date': row[10],
            'trial_id': row[11],
            'granted_by': row[12],
            'trial_type': row[13],
            'status': row[14],
            'extension_count': row[15],
            'time_remaining_hours': time_remaining,
            'is_expired': trial_end and datetime.now() >= trial_end if trial_end else False
        })
    
    conn.close()
    return jsonify({'trials': trials})

@app.route('/api/extend-trial/<client_id>', methods=['POST'])
def extend_trial_api(client_id):
    """Extend a trial period"""
    try:
        data = request.get_json()
        additional_hours = int(data.get('additional_hours', 24))
        extended_by = data.get('extended_by', 'admin')
        
        if additional_hours < 1 or additional_hours > 8760:
            return jsonify({'error': 'Additional hours must be between 1 and 8760'}), 400
        
        result = extend_trial(client_id, additional_hours, extended_by)
        
        if result['success']:
            return jsonify(result)
        else:
            return jsonify({'error': result['error']}), 500
            
    except Exception as e:
        return jsonify({'error': f'Failed to extend trial: {str(e)}'}), 500

@app.route('/api/convert-trial/<client_id>', methods=['POST'])
def convert_trial_api(client_id):
    """Convert trial to full account"""
    try:
        data = request.get_json()
        plan_type = data.get('plan_type', 'basic')
        
        result = convert_trial_to_full(client_id, plan_type)
        
        if result['success']:
            return jsonify(result)
        else:
            return jsonify({'error': result['error']}), 500
            
    except Exception as e:
        return jsonify({'error': f'Failed to convert trial: {str(e)}'}), 500

@app.route('/api/restrict-trial/<client_id>', methods=['POST'])
def restrict_trial_api(client_id):
    """Manually restrict a trial"""
    try:
        conn = sqlite3.connect('client_management.db')
        cursor = conn.cursor()
        
        # Restrict client access
        cursor.execute('''
            UPDATE clients 
            SET subscription_status = 'trial_expired', auto_restricted_at = ?
            WHERE client_id = ? AND account_type = 'trial'
        ''', (datetime.now().isoformat(), client_id))
        
        # Restrict all users
        cursor.execute('''
            UPDATE client_users 
            SET status = 'trial_expired', trial_restricted = 1
            WHERE client_id = ?
        ''', (client_id,))
        
        # Deactivate sessions
        cursor.execute('''
            UPDATE user_sessions 
            SET is_active = 0 
            WHERE client_id = ?
        ''', (client_id,))
        
        # Update trial management
        cursor.execute('''
            UPDATE trial_management 
            SET status = 'manually_restricted', auto_restricted_at = ?
            WHERE client_id = ?
        ''', (datetime.now().isoformat(), client_id))
        
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'message': 'Trial access restricted successfully'})
        
    except Exception as e:
        return jsonify({'error': f'Failed to restrict trial: {str(e)}'}), 500

@app.route('/api/visitors/<client_id>')
def get_visitors(client_id):
    """Get visitor data for a client with pagination"""
    try:
        page = int(request.args.get('page', 1))
        per_page = 50  # 50 visitors per page
        offset = (page - 1) * per_page
        
        conn = sqlite3.connect('client_management.db')
        cursor = conn.cursor()
        
        # Get total count
        cursor.execute('''
            SELECT COUNT(*) FROM visitor_investigations 
            WHERE client_id = ? AND status = 'active'
        ''', (client_id,))
        total_visitors = cursor.fetchone()[0]
        
        # Get visitors with pagination, sorted by interest level and time on site
        cursor.execute('''
            SELECT visitor_id, ip_address, country, city, device_type, browser,
                   current_page, pages_visited, time_on_site, interest_level,
                   contact_info, company_info, session_start, last_activity
            FROM visitor_investigations 
            WHERE client_id = ? AND status = 'active'
            ORDER BY 
                CASE interest_level 
                    WHEN 'high' THEN 1 
                    WHEN 'medium' THEN 2 
                    WHEN 'low' THEN 3 
                END,
                time_on_site DESC
            LIMIT ? OFFSET ?
        ''', (client_id, per_page, offset))
        
        visitors = []
        for row in cursor.fetchall():
            pages_visited = json.loads(row[7]) if row[7] else []
            contact_info = json.loads(row[10]) if row[10] else {}
            company_info = json.loads(row[11]) if row[11] else {}
            
            visitors.append({
                'visitor_id': row[0],
                'ip_address': row[1],
                'country': row[2],
                'city': row[3],
                'device_type': row[4],
                'browser': row[5],
                'current_page': row[6],
                'pages_visited': pages_visited,
                'time_on_site': row[8],
                'interest_level': row[9],
                'contact_info': contact_info,
                'company_info': company_info,
                'session_start': row[12],
                'last_activity': row[13]
            })
        
        conn.close()
        
        total_pages = (total_visitors + per_page - 1) // per_page
        
        return jsonify({
            'visitors': visitors,
            'pagination': {
                'current_page': page,
                'total_pages': total_pages,
                'total_visitors': total_visitors,
                'per_page': per_page,
                'has_next': page < total_pages,
                'has_prev': page > 1
            }
        })
        
    except Exception as e:
        return jsonify({'error': f'Failed to get visitors: {str(e)}'}), 500

@app.route('/api/export-visitors/<client_id>/<format>')
def export_visitors(client_id, format):
    """Export visitor data as CSV or Excel"""
    try:
        conn = sqlite3.connect('client_management.db')
        cursor = conn.cursor()
        
        cursor.execute('''
            SELECT visitor_id, ip_address, country, city, device_type, browser,
                   current_page, pages_visited, time_on_site, interest_level,
                   contact_info, company_info, session_start, last_activity
            FROM visitor_investigations 
            WHERE client_id = ? AND status = 'active'
            ORDER BY 
                CASE interest_level 
                    WHEN 'high' THEN 1 
                    WHEN 'medium' THEN 2 
                    WHEN 'low' THEN 3 
                END,
                time_on_site DESC
        ''', (client_id,))
        
        visitors = cursor.fetchall()
        conn.close()
        
        if format.lower() == 'csv':
            output = io.StringIO()
            writer = csv.writer(output)
            
            # Write header
            writer.writerow([
                'Visitor ID', 'IP Address', 'Country', 'City', 'Device', 'Browser',
                'Current Page', 'Pages Visited', 'Time on Site (seconds)', 'Interest Level',
                'Email', 'Phone', 'Name', 'Company', 'Session Start', 'Last Activity'
            ])
            
            # Write data
            for visitor in visitors:
                pages_visited = json.loads(visitor[7]) if visitor[7] else []
                contact_info = json.loads(visitor[10]) if visitor[10] else {}
                company_info = json.loads(visitor[11]) if visitor[11] else {}
                
                writer.writerow([
                    visitor[0], visitor[1], visitor[2], visitor[3], visitor[4], visitor[5],
                    visitor[6], '; '.join(pages_visited), visitor[8], visitor[9],
                    contact_info.get('email', ''), contact_info.get('phone', ''),
                    contact_info.get('name', ''), company_info.get('company', ''),
                    visitor[12], visitor[13]
                ])
            
            output.seek(0)
            return send_file(
                io.BytesIO(output.getvalue().encode()),
                mimetype='text/csv',
                as_attachment=True,
                download_name=f'visitors_{client_id}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
            )
            
        elif format.lower() == 'excel':
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = 'Visitors'
            
            # Write header
            headers = [
                'Visitor ID', 'IP Address', 'Country', 'City', 'Device', 'Browser',
                'Current Page', 'Pages Visited', 'Time on Site (seconds)', 'Interest Level',
                'Email', 'Phone', 'Name', 'Company', 'Session Start', 'Last Activity'
            ]
            
            for col, header in enumerate(headers, 1):
                cell = worksheet.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
            
            # Write data
            for row_num, visitor in enumerate(visitors, 2):
                pages_visited = json.loads(visitor[7]) if visitor[7] else []
                contact_info = json.loads(visitor[10]) if visitor[10] else {}
                company_info = json.loads(visitor[11]) if visitor[11] else {}
                
                data = [
                    visitor[0], visitor[1], visitor[2], visitor[3], visitor[4], visitor[5],
                    visitor[6], '; '.join(pages_visited), visitor[8], visitor[9],
                    contact_info.get('email', ''), contact_info.get('phone', ''),
                    contact_info.get('name', ''), company_info.get('company', ''),
                    visitor[12], visitor[13]
                ]
                
                for col, value in enumerate(data, 1):
                    worksheet.cell(row=row_num, column=col, value=value)
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            output = io.BytesIO()
            workbook.save(output)
            output.seek(0)
            
            return send_file(
                output,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=f'visitors_{client_id}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            )
        
        else:
            return jsonify({'error': 'Invalid format. Use csv or excel'}), 400
            
    except Exception as e:
        return jsonify({'error': f'Failed to export visitors: {str(e)}'}), 500

# Background task to check expired trials
def run_trial_monitor():
    """Background task to monitor and restrict expired trials"""
    while True:
        try:
            restricted_count = check_and_restrict_expired_trials()
            if restricted_count > 0:
                print(f"Automatically restricted {restricted_count} expired trials")
        except Exception as e:
            print(f"Error in trial monitor: {e}")
        
        time.sleep(3600)  # Check every hour

# Start background trial monitor
trial_monitor_thread = threading.Thread(target=run_trial_monitor, daemon=True)
trial_monitor_thread.start()

@app.route('/health')
def health_check():
    """Health check endpoint with trial monitoring"""
    # Run trial check on health check
    restricted_count = check_and_restrict_expired_trials()
    
    return jsonify({
        'status': 'healthy', 
        'timestamp': datetime.now().isoformat(),
        'trials_restricted': restricted_count
    })

if __name__ == '__main__':
    print("🚀 Starting Enhanced Trial Management System")
    print("📝 Main App: http://localhost:5000/")
    print("⚙️ Admin Dashboard: http://localhost:5000/admin")
    print("🕐 Trial Management: http://localhost:5000/admin/trials")
    print("👥 User Management: http://localhost:5000/admin/users")
    print("🔄 Features: Flexible trials, automatic restrictions, pagination, export")
    app.run(host='0.0.0.0', port=5000, debug=T
